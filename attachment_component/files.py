import os
import PyPDF2
import pandas as pd
from PIL import Image
import pytesseract
from openpyxl import load_workbook
from docx import Document

if not os.path.exists('logs'):
    os.makedirs('logs')

def log_error(message):
    with open('logs/files.txt', 'a') as log_file:
        log_file.write(message + "\n")

def extract_text_from_pdf(file_path):
    try:
        text = ""
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        log_error(f"Error extracting text from PDF file '{file_path}': {str(e)}")
        return None

def extract_text_from_image(file_path):
    try:
        image = Image.open(file_path)
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        log_error(f"Error extracting text from image file '{file_path}': {str(e)}")
        return None

def extract_text_from_csv(file_path):
    try:
        df = pd.read_csv(file_path)
        return df.to_string(index=False)
    except Exception as e:
        log_error(f"Error extracting text from CSV file '{file_path}': {str(e)}")
        return None

def extract_text_from_excel(file_path):
    try:
        text = ""
        workbook = load_workbook(file_path)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                text += " ".join(str(cell) for cell in row if cell is not None) + "\n"
        return text
    except Exception as e:
        log_error(f"Error extracting text from Excel file '{file_path}': {str(e)}")
        return None

def extract_text_from_docx(file_path):
    try:
        doc = Document(file_path)
        text = "\n".join([para.text for para in doc.paragraphs])
        return text
    except Exception as e:
        log_error(f"Error extracting text from DOCX file '{file_path}': {str(e)}")
        return None

def extract_text_from_txt(file_path):
    try:
        with open(file_path, 'r') as file:
            return file.read()
    except Exception as e:
        log_error(f"Error extracting text from TXT file '{file_path}': {str(e)}")
        return None

def extract_text_from_file(file_path):
    extension = os.path.splitext(file_path)[1].lower()
    if extension == '.pdf':
        return extract_text_from_pdf(file_path)
    elif extension in ['.png', '.jpg', '.jpeg']:
        return extract_text_from_image(file_path)
    elif extension == '.csv':
        return extract_text_from_csv(file_path)
    elif extension in ['.xls', '.xlsx']:
        return extract_text_from_excel(file_path)
    elif extension == '.docx':
        return extract_text_from_docx(file_path)
    elif extension == '.txt':
        return extract_text_from_txt(file_path)
    else:
        log_error(f"Unsupported file extension '{extension}' for file '{file_path}'.")
        return None
